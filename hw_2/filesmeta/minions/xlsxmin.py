'''Minion get xlsx files metadata'''

from openpyxl import load_workbook

class XlsxMinion():
    '''Minion get xlsx files metadata'''
    def __init__(self):
        self.columns = ['MS_Office_Category', 'MS_Office_Contentstatus', 'MS_Office_Creator',
                        'MS_Office_Description', 'MS_Office_Identifier', 'MS_Office_Keywords',
                        'MS_Office_Language', 'MS_Office_Last_modified_by', 'MS_Office_Revision',
                        'MS_Office_Subject', 'MS_Office_Title', 'MS_Office_Version',
                        'MS_Office_Sheet_Count']
        self.ex = ['xlsx', 'xlsm', 'xltx', 'xltm']
        self._metadata = ()

    def get_meta_inf(self, file):
        '''get metadata'''
        try:
            workbook = load_workbook(file)

            self._metadata = (workbook.properties.category,
                                workbook.properties.contentStatus,
                                workbook.properties.creator,
                                workbook.properties.description,
                                workbook.properties.identifier,
                                workbook.properties.keywords,
                                workbook.properties.language,
                                workbook.properties.last_modified_by,
                                workbook.properties.revision,
                                workbook.properties.subject,
                                workbook.properties.title,
                                workbook.properties.version, 
                                len(workbook.sheetnames))

            return dict(zip(self.columns, self._metadata))
        
        except Exception as e:
            print(e)